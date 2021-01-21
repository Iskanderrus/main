import sys
import os
import openpyxl
import pandas as pd

file = 'Expences_2020.xlsx'
data = pd.ExcelFile(file)
print(data.sheet_names) #this returns the all the sheets in the excel file

df = data.parse("1609270171129")

ps = openpyxl.load_workbook(file)
sheet = ps["1609270171129"]
sheet.max_row
print(sheet.max_row)

for row in range(2, sheet.max_row + 1):
    multipled = df['Сумма'] * 2
    df['Multipled'] = multipled
print(df.tail(10))
df.to_excel("output.xlsx")